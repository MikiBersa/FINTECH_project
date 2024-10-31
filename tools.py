# DOWNLOAD FILE EXCEL
import os
from openpyxl import Workbook, load_workbook
import json
import pandas as pd

# function that creates a new sheet if it does not already exist
def add_sheet_to_excel(filename, sheet_name):
    # Check if the file exists
    if os.path.exists(filename):
        # Load the existing workbook
        workbook = load_workbook(filename)
        print(f"The file '{filename}' exists. I add a new sheet.")
    else:
        # Create a new workbook
        workbook = Workbook()
        print(f"The file '{filename}' does not exist. I create a new file.")

    # Add a new sheet
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
        print(f"Sheet '{sheet_name}' added.")
    else:
        print(f"Sheet '{sheet_name}' already exists.")

    # Save the workbook
    workbook.save(filename)
    print(f"File saved as '{filename}'.")

# function that adds the data in the sheet with the name sheet_name
def add_data_to_sheet(filename, sheet_name, data):
    # Upload the workbook
    workbook = load_workbook(filename)

    # Check if the sheet exists
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # I DELETE THE PREVIOUS VALUES I CLEAN THE SHEET
        if sheet.max_row > 0:
            sheet.delete_rows(0, sheet.max_row)
    else:
        print(f"Sheet '{sheet_name}' does not exist. I add a new sheet.")
        sheet = workbook.create_sheet(title=sheet_name)

    sheet.append(data.columns.tolist())

    data_values = data.values.tolist()

    # Add data to the sheet
    for row in data_values:
        sheet.append(row)

    # Save the workbook
    workbook.save(filename)
    print(f"Data added and file saved as '{filename}'.")

# This function returns the score of the model identified by the dataset and the model name
def extract_data_model(file_path, dataset, model_name, what):
    with open(file_path, 'r') as file:
        settings = json.load(file)

    for set in settings:
        if set['dataset'] == dataset:
            if what == "params":
                for model in set['settings']:
                    if model["model"] == model_name:
                        return model["params"]
            elif what == "scores":
                for model in set['scores']:
                    if model["model"] == model_name:
                        ritorno = {}

                        ritorno["accuracy"] = model["accuracy"]
                        ritorno["f1_score"]  = model["f1_score"]
                        ritorno["recall"] = model["recall"]
                        ritorno["precision"] = model["precision"]
                        ritorno["roc_auc"] = model["roc_auc"]

                        return ritorno

# This function returns the score of the model identified by only by the model name
# So it iterates over the different datasets
def extract_scores_model_dataset(file_path, model_name):
    ritorno = {}
    ritorno["dataset"] = []
    ritorno["f1_score"] = []

    with open(file_path, 'r') as file:
        settings = json.load(file)

    for set in settings:
        scores = set["scores"]
        ritorno["dataset"].append(set["dataset"])

        for score in scores:
            if score["model"] == model_name:
                ritorno["f1_score"].append(score["f1_score"])


    return ritorno

# This function saves the column of the dataset in the settings file
def save_columns_settings(file_path, columns, dataset_name):
    try:
        # 1. Read the JSON file
        with open(file_path, 'r') as f:
            data = json.load(f)

        # 2. Edit the content
        for dataset in data:
            if dataset["dataset"] == dataset_name:
                dataset["columns"] = columns
                break

        # 3. Save the modified JSON file
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)

    except FileNotFoundError:
        print("The file was not found.")
    except json.JSONDecodeError:
        print("Error decoding JSON file.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return

# This function saves the weights of the feature that are used in the model_name
# so it saves as a feature selection
def save_feature_settings(file_path, feature, dataset_name, model_name):
    try:
        # Read the JSON file
        with open(file_path, 'r') as f:
            data = json.load(f)

        for dataset in data:
            if dataset["dataset"] == dataset_name:
                models = dataset["scores"]
                for model in models:
                    if model["model"] == model_name:
                        model["features"] = feature.tolist()
                        break

        # Save the modified JSON file
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)

    except FileNotFoundError:
        print("The file was not found.")
    except json.JSONDecodeError:
        print("Error decoding JSON file.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return

# this function extracts the information about the feature of a model over the datasets
def extract_feature_model_dataset(file_path, model_name):
    ritorno = []

    with open(file_path, 'r') as file:
        settings = json.load(file)

    for set in settings:
        info = {}

        info["dataset"] = set["dataset"]
        info["columns"] = []
        info["columns"] = set["columns"]
        info["weights"] = []

        scores = set["scores"]

        for score in scores:
            if score["model"] == model_name:
                info["weights"] = score["features"]
                break

        ritorno.append(info)

    return ritorno


# this function extractes the score of an model over the dataset
# in particular it extracts the f1_score
def extract_scores_model_dataset(file_path, model_name):
    ritorno = {}
    ritorno["dataset"] = []
    ritorno["f1_score"] = []

    with open(file_path, 'r') as file:
        settings = json.load(file)

    for set in settings:
        scores = set["scores"]
        ritorno["dataset"].append(set["dataset"])

        for score in scores:
            if score["model"] == model_name:
                ritorno["f1_score"].append(score["f1_score"])
    return ritorno

# This function extracts the scores of a specific voting model over the datasets
def extract_scores_model_voting(file_path, voting_model_name, valutation_score):
    ritorno = {}
    ritorno["dataset"] = []
    ritorno["f1_score"] = []

    with open(file_path, 'r') as file:
        settings = json.load(file)

    for set in settings:
        voting = set["voting"]
        ritorno["dataset"].append(set["dataset"])

        for vote in voting:
            if vote["model"] == voting_model_name:
                ritorno["f1_score"].append(vote[valutation_score])
    return ritorno

# This function loads the list of the models that we have used
def get_models(file_path):
    ritorno_models = []
    ritorno_dataset = []
    ritorno_voting = []

    with open(file_path, 'r') as file:
        settings = json.load(file)

    set = settings[0]
    scores = set["scores"]
    votings = set["voting"]

    for score in scores:
        ritorno_models.append(score["model"])

    for dataset in settings:
        ritorno_dataset.append(dataset["dataset"])

    for voting in votings:
        ritorno_voting.append(voting["model"])

    return ritorno_models, ritorno_dataset, ritorno_voting

# This function saves the score of a specific model in a specific dataset
def save_score_settings(file_settings, score, dataset_name, model_name):
    try:
        # 1. Read the JSON file
        with open(file_settings, 'r') as f:
            data = json.load(f)

        # 2. Edit the content
        for dataset in data:
            if dataset["dataset"] == dataset_name:
                for model in dataset["scores"]:
                    if model["model"] == model_name:
                        model["accuracy"] = score["accuracy"]
                        model["f1_score"] = score["f1_score"]
                        model["recall"] = score["recall"]
                        model["precision"] = score["precision"]
                        model["roc_auc"] = score["roc_auc"]

                        break

        # 3. Save the modified JSON file
        with open(file_settings, 'w') as f:
            json.dump(data, f, indent=4)

    except FileNotFoundError:
        print("The file was not found.")
    except json.JSONDecodeError:
        print("Error decoding JSON file.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return

# This function saves the scores of a specific voting model in a specific dataset
def save_voting_scores_settings(file_settings, score, dataset_name, model_name):
    try:
        # 1. Read the JSON file
        with open(file_settings, 'r') as f:
            data = json.load(f)

        # 2. Edit the content
        for dataset in data:
            if dataset["dataset"] == dataset_name:
                for model in dataset["voting"]:
                    if model["model"] == model_name:
                        model["accuracy"] = score["accuracy"]
                        model["f1_score"] = float(score["f1_score"])
                        model["recall"] = score["recall"]
                        model["precision"] = score["precision"]
                        model["roc_auc"] = score["roc_auc"]
                        model["loss"] = score["loss"]
                        model["accuracy_deep"] = score["accuracy_deep"]

                        break

        # 3.  Save the modified JSON file
        with open(file_settings, 'w') as f:
            json.dump(data, f, indent=4)

    except FileNotFoundError:
        print("The file was not found.")
    except json.JSONDecodeError:
        print("Error decoding JSON file.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return

# This function is a support function in order to create the mix model
# So it returns le list of setting of every model of a specific dataset
def load_params_mix(file_path, dataset):
    with open(file_path, 'r') as file:
        settings = json.load(file)

    structure = {}

    for set in settings:
        if set["dataset"] == dataset:

            for model in set["settings"]:
                structure[model["model"]] = model["params"]["params"]

            break

    return structure

# This function save the hedge_ratio and intercept in the settings file in the dataset STATIONARITY
def seve_hedge_ratio(file_path, hedge_ratio, intercept):
    try:
        # 1. Read the JSON file
        with open(file_path, 'r') as f:
            data = json.load(f)

        # 2. Edit the content
        for dataset in data:
            if dataset["dataset"] == "STATIONARITY":
                dataset["hedge_ratio"] = hedge_ratio
                dataset["intercept"] = intercept
                break

        # 3.  Save the modified JSON file
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)

    except FileNotFoundError:
        print("The file was not found.")
    except json.JSONDecodeError:
        print("Error decoding JSON file.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return

# This function extrapolates the cointegrated pairs used in the datasets
def extract_pairs(file_path, dataset):
    cointegrated_pairs = []

    with open(file_path, 'r') as f:
                settings = json.load(f)

    for setting in settings:
        if setting["dataset"] == dataset:
            i = 0
            for pair in setting["columns"]:
                asset_split = pair.split("_")
                cointegrated_pairs.append((asset_split[1], asset_split[2], i))
                i += 1

    return cointegrated_pairs

# This function processes the input data so that each model can do
# the classification by pretending to have its own dataset on which they trained
def data_preparation_prediction(file_path, data, list_dataset):
    ritorno = {}
    with open(file_path, 'r') as f:
            settings = json.load(f)

    for dataset in list_dataset:
        if dataset == "SPREAD":
            cointegrated_pairs = extract_pairs(file_path, "SPREAD")
            # In DATA we calculate the spread between the asset in pairs
            cointegrated_data = pd.concat([data[col1] - data[col2] for col1, col2, i in cointegrated_pairs], axis=1)
            cointegrated_data.columns = [f"Coint_{col1}_{col2}" for col1, col2, i in cointegrated_pairs]
            ritorno["SPREAD"] = cointegrated_data

        elif dataset == "STATIONARITY":
            hedge_ratio = []
            intercept = []

            for setting in settings:
                if setting["dataset"] == dataset:
                    hedge_ratio = setting["hedge_ratio"]
                    intercept = setting["intercept"]

            # In STATIONARITY we calculate the spread between the linear combiantion between the asset in pairs
            cointegrated_pairs = extract_pairs(file_path, "STATIONARITY")
            linear_comb = pd.concat([data[col1] - (hedge_ratio[i] * data[col2] + intercept[i]) for col1, col2, i in cointegrated_pairs], axis=1)
            linear_comb.columns = [f"Coint_{col1}_{col2}" for col1, col2, i in cointegrated_pairs]
            ritorno["STATIONARITY"] = linear_comb

    return ritorno